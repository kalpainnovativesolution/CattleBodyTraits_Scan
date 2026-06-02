"""
cattle_traits_app.py  ─  Streamlit Web App for Cattle Trait Measurement
========================================================================
Side-view  : ResNet101 KeypointRCNN (15 keypoints)
Rear-view  : YOLOv8-pose (16 keypoints)
Scale      : SAM2 sticker click (15 cm square) via canvas widget

CHANGELOG:
- Added foot_angle (acute angle: pastern_hoof_junction → heel_bulb → hoof_tip)
- Added foot_angle_score (1-9 scale based on angle ranges)
- Both stored in side measurements and logs
- Full mobile-responsive layout (10-trait grid, auto-close sidebar on nav)
- Fixed light-mode image display glitch
"""

import math, os, json, pickle, io, copy
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
import streamlit as st
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from model_files import ensure_model_file, model_file_available

# ── optional heavy imports ────────────────────────────────────────────────────
try:
    import torch
    import torchvision.transforms.functional as TF
    from torchvision.models.detection import KeypointRCNN
    from torchvision.models.detection.backbone_utils import resnet_fpn_backbone
    TORCH_OK = True
except ImportError:
    TORCH_OK = False
    TORCH_IMPORT_ERROR = "PyTorch or torchvision is not installed."
except Exception as e:
    TORCH_OK = False
    TORCH_IMPORT_ERROR = f"{type(e).__name__}: {e}"

try:
    from ultralytics import YOLO, SAM
    YOLO_OK = True
except ImportError:
    YOLO_OK = False

try:
    from streamlit_image_coordinates import streamlit_image_coordinates
    COORDS_OK = True
except ImportError:
    COORDS_OK = False

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CattleScan",
    page_icon="🐄",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS  — full mobile-responsive overhaul
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Global ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
    color: white;
}
[data-testid="stSidebar"] * { color: white !important; }
[data-testid="stSidebar"] .stRadio label { color: white !important; }

/* ── Header banner ── */
.app-header {
    background: linear-gradient(135deg, #1a3a5c 0%, #2d6a4f 100%);
    border-radius: 16px;
    padding: 28px 36px;
    margin-bottom: 24px;
    display: flex; align-items: center; gap: 20px;
    box-shadow: 0 4px 24px rgba(0,0,0,0.18);
}
.app-header h1 { color: white; margin: 0; font-size: 2rem; font-weight: 700; }
.app-header p  { color: #a8d8b0; margin: 4px 0 0; font-size: 0.95rem; }
.header-icon   { font-size: 3.2rem; line-height: 1; }

/* ── Trait grid — responsive ── */
.trait-grid {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 14px;
    margin: 20px 0;
}
/* Tablet */
@media (max-width: 1100px) {
    .trait-grid { grid-template-columns: repeat(3, 1fr); }
}
/* Mobile */
@media (max-width: 640px) {
    .trait-grid { grid-template-columns: repeat(2, 1fr); gap: 10px; }
    .app-header { padding: 16px 18px; }
    .app-header h1 { font-size: 1.4rem; }
    .header-icon { font-size: 2.2rem; }
}

.trait-card {
    background: linear-gradient(135deg, #1a3a5c, #1e4d6b);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 14px;
    padding: 16px 12px;
    text-align: center;
    box-shadow: 0 2px 12px rgba(0,0,0,0.15);
    transition: transform 0.2s;
    min-height: 90px;
}
.trait-card:hover { transform: translateY(-3px); }
.trait-card .tc-label {
    color: #8ecae6; font-size: 0.72rem; font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 6px;
    line-height: 1.3;
}
.trait-card .tc-value {
    color: #ffffff; font-size: 1.45rem; font-weight: 700;
}
.trait-card .tc-unit  { color: #90caf9; font-size: 0.76rem; margin-top: 2px; }

/* ── Section headings ── */
.section-head {
    background: linear-gradient(90deg, #1a3a5c, transparent);
    border-left: 4px solid #40916c;
    padding: 8px 16px; border-radius: 4px;
    color: #e0f2fe; font-weight: 600; font-size: 1.05rem;
    margin: 20px 0 12px;
}

/* ── Sticker prompt box ── */
.sticker-box {
    background: linear-gradient(135deg, #0a2342, #0d3460);
    border: 2px dashed #40916c;
    border-radius: 12px; padding: 18px;
    text-align: center; color: #a8d8b0;
    font-weight: 500; margin: 12px 0;
}

/* ── Log table ── */
.log-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
.log-table th {
    background: #1a3a5c; color: #8ecae6;
    padding: 8px 10px; text-align: left;
    font-weight: 600; font-size: 0.75rem; text-transform: uppercase;
    letter-spacing: 0.05em;
}
.log-table td { padding: 7px 10px; border-bottom: 1px solid rgba(255,255,255,0.06); color: #5299c9; }
.log-table tr:hover td { background: rgba(64,145,108,0.08); }

/* ── Buttons ── */
div[data-testid="stButton"] > button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    letter-spacing: 0.03em !important;
}

/* ── CRITICAL FIX: Images always visible in both light and dark mode ── */
[data-testid="stImage"] {
    background: transparent !important;
}
[data-testid="stImage"] img,
[data-testid="stImage"] > div > img {
    max-width: 100% !important;
    max-height: 460px !important;
    width: auto !important;
    height: auto !important;
    object-fit: contain !important;
    border-radius: 8px;
    display: block !important;
    /* Force visibility in ALL themes — fixes light-mode invisible image bug */
    opacity: 1 !important;
    visibility: visible !important;
    background-color: transparent !important;
    -webkit-filter: none !important;
    filter: none !important;
}
/* Also force the stImage wrapper to be visible */
[data-testid="stImage"] > div {
    background: transparent !important;
    display: block !important;
}

/* streamlit-image-coordinates */
iframe {
    max-width: 100% !important;
    width: 100% !important;
}
[data-testid="stCustomComponentV1"] {
    max-width: 100% !important;
    overflow: hidden !important;
}

/* make every column not overflow horizontally */
[data-testid="stHorizontalBlock"] > div {
    overflow: hidden !important;
    min-width: 0 !important;
}

/* ── Info / warning boxes ── */
.info-box {
    background: rgba(64,145,108,0.12);
    border: 1px solid #40916c;
    border-radius: 10px; padding: 12px 16px;
    color: #a8d8b0; margin: 10px 0;
}
.warn-box {
    background: rgba(230,162,0,0.12);
    border: 1px solid #e6a200;
    border-radius: 10px; padding: 12px 16px;
    color: #ffd60a; margin: 10px 0;
}

/* ── Sidebar nav ── */
[data-testid="stSidebar"] .stRadio > div { gap: 4px; }
[data-testid="stSidebar"] .stRadio label {
    cursor: pointer !important;
    padding: 6px 10px !important;
    border-radius: 6px !important;
    transition: background 0.15s !important;
}
[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(64,145,108,0.25) !important;
}

/* ── Mobile layout fixes ── */
@media (max-width: 768px) {
    /* Stack side+rear columns vertically on mobile */
    [data-testid="stHorizontalBlock"] { flex-direction: column !important; }
    [data-testid="stHorizontalBlock"] > div { width: 100% !important; flex: none !important; }

    /* Ensure uploaded images are ALWAYS visible on mobile light mode */
    [data-testid="stImage"],
    [data-testid="stImage"] > div,
    [data-testid="stImage"] img {
        display: block !important;
        opacity: 1 !important;
        visibility: visible !important;
        max-width: 100% !important;
        height: auto !important;
        background: transparent !important;
    }

    /* Ensure file uploader is accessible */
    [data-testid="stFileUploader"] {
        width: 100% !important;
    }

    /* Camera input full width */
    [data-testid="stCameraInput"] {
        width: 100% !important;
    }
}

/* ── Force close sidebar on mobile via JS injection (applied via component) ── */
</style>

<script>
// Auto-close sidebar on mobile when a nav radio button is clicked
(function() {
    function closeSidebarOnMobile() {
        if (window.innerWidth <= 768) {
            // Find Streamlit's sidebar close button and click it
            setTimeout(function() {
                var closeBtn = document.querySelector('[data-testid="stSidebarCollapseButton"]');
                if (closeBtn) closeBtn.click();
                // Also try the hamburger toggle
                var toggle = document.querySelector('[aria-label="Close sidebar"]');
                if (toggle) toggle.click();
            }, 200);
        }
    }

    // Watch for radio button changes in sidebar
    document.addEventListener('change', function(e) {
        if (e.target && e.target.closest('[data-testid="stSidebar"]')) {
            closeSidebarOnMobile();
        }
    });

    // Also watch for clicks on sidebar labels
    document.addEventListener('click', function(e) {
        if (e.target && e.target.closest('[data-testid="stSidebar"] .stRadio')) {
            closeSidebarOnMobile();
        }
    });
})();
</script>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
STICKER_REAL_CM = 15.0
NUM_CLASSES     = 2
NUM_KEYPOINTS   = 15
SCORE_THRESH    = 0.5
LOG_FILE        = "cattle_logs.json"
EXCEL_FILE      = "cattle_logs.xlsx"
OUTPUT_DIR      = Path("cattle_output")
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Side-view keypoints ───────────────────────────────────────────────────────
SIDE_KP_NAMES = [
    "wither","foot","chest_top","chest_bottom","body_top","body_bottom",
    "hip_bone","pin_bone","shoulder_bone","stifle_thigh","hock",
    "pastern","hoof_tip","heel_bulb","pastern_hoof_junction",
]
SKP = {n: i for i, n in enumerate(SIDE_KP_NAMES)}

SIDE_KP_COLORS = [
    (0,255,0),(0,200,255),(255,100,0),(255,0,100),(0,100,255),
    (0,0,255),(200,0,200),(0,255,200),(128,255,0),(255,128,0),
    (0,128,255),(255,0,200),(200,255,0),(128,0,255),(255,255,0),
]

# ── Rear-view keypoints ───────────────────────────────────────────────────────
REAR_KP_NAMES = [
    "rump_left","rump_right","left_thigh","right_thigh",
    "left_hock","right_hock","hoof_left","hoof_right",
    "abdomen_left","abdomen_right","belly_center","tailtop",
    "bellytopleft","bellytopright","uddertop","udderbottom",
]
RKP = {n: i for i, n in enumerate(REAR_KP_NAMES)}

REAR_MEASUREMENTS = [
    ("tailtop",       "bellytopright", "M01_tailtop-bellytopright"),
    ("tailtop",       "bellytopleft",  "M02_tailtop-bellytopleft"),
    ("rump_left",     "rump_right",    "M03_rump_left-rump_right"),
    ("bellytopleft",  "abdomen_left",  "M04_bellytopleft-abdomen_left"),
    ("bellytopright", "abdomen_right", "M05_bellytopright-abdomen_right"),
    ("rump_left",     "abdomen_left",  "M06_rump_left-abdomen_left"),
    ("rump_right",    "abdomen_right", "M07_rump_right-abdomen_right"),
    ("tailtop",       "uddertop",      "M08_tailtop-uddertop"),
    ("uddertop",      "udderbottom",   "M09_uddertop-udderbottom"),
    ("abdomen_left",  "abdomen_right", "M10_abdomen_left-abdomen_right"),
    ("udderbottom",   "belly_center",  "M11_udderbottom-belly_center"),
    ("left_thigh",    "belly_center",  "M12_left_thigh-belly_center"),
    ("right_thigh",   "belly_center",  "M13_right_thigh-belly_center"),
    ("abdomen_left",  "left_thigh",    "M14_abdomen_left-left_thigh"),
    ("abdomen_right", "right_thigh",   "M15_abdomen_right-right_thigh"),
    ("left_thigh",    "left_hock",     "M16_left_thigh-left_hock"),
    ("right_thigh",   "right_hock",    "M17_right_thigh-right_hock"),
    ("left_hock",     "hoof_left",     "M18_left_hock-hoof_left"),
    ("right_hock",    "hoof_right",    "M19_right_hock-hoof_right"),
    ("left_hock",     "right_hock",    "M20_left_hock-right_hock"),
    ("left_thigh",    "right_thigh",   "M21_thigh_width"),
]

REAR_SKELETON = [
    ("rump_left","rump_right"),("rump_left","abdomen_left"),
    ("rump_right","abdomen_right"),("abdomen_left","abdomen_right"),
    ("tailtop","bellytopright"),("tailtop","bellytopleft"),
    ("bellytopleft","abdomen_left"),("bellytopright","abdomen_right"),
    ("tailtop","uddertop"),("uddertop","udderbottom"),
    ("udderbottom","belly_center"),("belly_center","left_thigh"),
    ("belly_center","right_thigh"),("abdomen_left","left_thigh"),
    ("abdomen_right","right_thigh"),("left_thigh","left_hock"),
    ("right_thigh","right_hock"),("left_hock","hoof_left"),
    ("right_hock","hoof_right"),("left_hock","right_hock"),
]

REAR_KP_COLORS = [
    (0,255,255),(0,200,255),(100,255,100),(60,200,100),
    (255,0,0),(200,0,0),(0,0,255),(0,0,180),
    (255,150,0),(255,100,0),(255,0,200),(0,255,0),
    (150,255,0),(200,255,0),(0,180,255),(0,120,255),
]

# ─────────────────────────────────────────────────────────────────────────────
# FOOT ANGLE SCORE TABLE
# ─────────────────────────────────────────────────────────────────────────────
FOOT_ANGLE_SCORE_TABLE = [
    (70, 76, 1),
    (63, 69, 2),
    (56, 62, 3),
    (49, 55, 4),
    (42, 48, 5),
    (35, 41, 6),
    (28, 34, 7),
    (21, 27, 8),
    (14, 20, 9),
]

def foot_angle_to_score(angle_deg):
    """Convert foot angle (degrees) to 1-9 score. Returns None if out of range."""
    if angle_deg is None:
        return None
    for lo, hi, score in FOOT_ANGLE_SCORE_TABLE:
        if lo <= angle_deg <= hi:
            return score
    # Out of defined range — clamp to nearest
    if angle_deg > 76:
        return 1
    if angle_deg < 14:
        return 9
    return None

# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────────────────────────
def init_state():
    defaults = {
        "tag": "",
        "nav": "Measure",
        # side
        "side_img_raw": None,
        "side_img_display": None,
        "side_sticker_done": False,
        "side_cm_per_px": None,
        "side_click_coords": None,
        "side_last_processed_click": None,
        "side_file_id": None,
        "side_results": None,
        "side_annotated": None,
        # rear
        "rear_img_raw": None,
        "rear_img_display": None,
        "rear_sticker_done": False,
        "rear_cm_per_px": None,
        "rear_click_coords": None,
        "rear_last_processed_click": None,
        "rear_file_id": None,
        "rear_results": None,
        "rear_annotated": None,
        # final
        "traits": None,
        "processing": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────────────────────────────────────
# MODEL LOADERS  (cached)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner="Loading Side-view model…")
def load_side_model(ckpt_path):
    if not TORCH_OK:
        return None
    ckpt_path = ensure_model_file(ckpt_path)
    backbone = resnet_fpn_backbone(backbone_name='resnet101', weights=None)
    model    = KeypointRCNN(backbone, num_classes=NUM_CLASSES,
                            num_keypoints=NUM_KEYPOINTS)
    device   = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    ckpt     = torch.load(ckpt_path, map_location=device)
    sd       = ckpt.get("model_state_dict", ckpt)
    model.load_state_dict(sd)
    model.to(device).eval()
    return model, device

@st.cache_resource(show_spinner="Loading Rear-view model…")
def load_rear_model(yolo_path):
    if not YOLO_OK:
        return None
    yolo_path = ensure_model_file(yolo_path)
    return YOLO(yolo_path)

@st.cache_resource(show_spinner="Loading SAM2…")
def load_sam2(sam2_path):
    if not YOLO_OK:
        return None
    try:
        sam2_path = ensure_model_file(sam2_path)
        return SAM(sam2_path)
    except Exception as e:
        st.warning(f"SAM2 load failed: {e}")
        return None

@st.cache_resource(show_spinner="Loading RF score model…")
def load_rf_model(pkl_path):
    try:
        pkl_path = ensure_model_file(pkl_path)
    except Exception:
        return None
    try:
        with open(pkl_path, "rb") as f:
            return pickle.load(f)
    except Exception:
        pass
    try:
        import joblib
        return joblib.load(pkl_path)
    except Exception:
        pass
    try:
        with open(pkl_path, "rb") as f:
            return pickle.load(f, encoding="latin1")
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────────────────
# GEOMETRY HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def dist2d(p1, p2):
    return math.hypot(p2[0]-p1[0], p2[1]-p1[1])

def px_cm(px, cmp):
    return px * cmp if (px is not None and cmp) else None

def angle_at_vertex(a, b, c):
    """Acute angle at point b formed by vectors b→a and b→c."""
    ba = np.array([a[0]-b[0], a[1]-b[1]], dtype=float)
    bc = np.array([c[0]-b[0], c[1]-b[1]], dtype=float)
    n  = np.linalg.norm(ba)*np.linalg.norm(bc)
    if n < 1e-9: return None
    cos_a = np.dot(ba, bc) / n
    return float(np.degrees(np.arccos(np.clip(cos_a, -1.0, 1.0))))

def acute_angle_with_horizontal(p_from, p_to):
    dx = p_to[0] - p_from[0]
    dy = p_to[1] - p_from[1]
    ang = abs(math.degrees(math.atan2(abs(dy), abs(dx))))
    return min(ang, 180-ang)

# ─────────────────────────────────────────────────────────────────────────────
# SAM2 STICKER SEGMENTATION
# ─────────────────────────────────────────────────────────────────────────────
def _order_points(pts):
    pts  = np.array(pts, dtype="float32")
    rect = np.zeros((4,2), dtype="float32")
    s    = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]; rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]; rect[3] = pts[np.argmax(diff)]
    return rect

def _contour_to_quad(cnt):
    peri = cv2.arcLength(cnt, True)
    for eps in [0.02,0.03,0.04,0.05,0.07,0.10]:
        ap = cv2.approxPolyDP(cnt, eps*peri, True)
        if len(ap)==4:
            return _order_points(ap.reshape(4,2))
    rect = cv2.minAreaRect(cnt)
    box  = cv2.boxPoints(rect)
    return _order_points(box)

def segment_sticker_sam2(img_bgr, click_xy, sam2_model):
    try:
        pts    = np.array([[click_xy[0], click_xy[1]]])
        labels = np.array([1])
        results = sam2_model(img_bgr, points=pts, labels=labels)
        if not results or results[0].masks is None:
            return None, None
        mask = results[0].masks.data[0].cpu().numpy()
        mask = (mask*255).astype(np.uint8)
        kern = np.ones((5,5), np.uint8)
        mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN,  kern)
        mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kern)
        cnts, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not cnts: return None, None
        best = max(cnts, key=cv2.contourArea)
        if cv2.contourArea(best) < 300: return None, None
        corners = _contour_to_quad(best)
        sides   = [np.linalg.norm(corners[i]-corners[(i+1)%4]) for i in range(4)]
        stk_px  = max(sides)
        cmp     = STICKER_REAL_CM / stk_px
        return cmp, corners
    except Exception as e:
        st.error(f"SAM2 error: {e}")
        return None, None

def fallback_cm_per_px(img_bgr, click_xy, radius_px=40):
    return STICKER_REAL_CM / (radius_px * 2), None

# ─────────────────────────────────────────────────────────────────────────────
# SIDE-VIEW INFERENCE
# ─────────────────────────────────────────────────────────────────────────────
def run_side_inference(img_bgr, model, device):
    from PIL import Image as PILImage
    rgb  = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil  = PILImage.fromarray(rgb)
    tens = TF.to_tensor(pil).to(device)
    with torch.no_grad():
        out = model([tens])[0]
    keep = out["scores"] >= SCORE_THRESH
    scores   = out["scores"][keep].cpu().numpy()
    kps_all  = out["keypoints"][keep].cpu().numpy()
    boxes  = out["boxes"][keep].cpu().numpy()
    if len(scores)==0: return None, None
    best = int(np.argmax(scores))
    return kps_all[best], float(scores[best]), boxes[best]

def compute_side_measurements(kps, cmp,bbox=None):
    """
    kps: (15,3) array of (x,y,vis)

    NEW: foot_angle — acute angle at heel_bulb (vertex) from:
         pastern_hoof_junction → heel_bulb → hoof_tip
    NEW: foot_angle_score — 1-9 score based on angle range table
    """
    def p(name):
        return kps[SKP[name]][:2]
    def vis(name):
        return float(kps[SKP[name]][2]) > 0

    def safe_dist(n1, n2):
        if vis(n1) and vis(n2):
            return dist2d(p(n1), p(n2))
        return None

    # distances in px
    body_length_px        = safe_dist("shoulder_bone", "pin_bone")
    body_height_px        = safe_dist("wither",        "foot")
    # BBox height: y2 − y1 of the detection bounding box
    bbox_height_px       = float(bbox[3] - bbox[1]) if bbox is not None else None
    hip_length_px         = safe_dist("hip_bone",      "pin_bone")
    chest_height_px       = safe_dist("chest_top",     "chest_bottom")
    linear_body_depth_px  = safe_dist("body_top",      "body_bottom")

    # Rump vertical (signed)
    rump_vertical_px = None
    if vis("hip_bone") and vis("pin_bone"):
        rump_vertical_px = float(p("pin_bone")[1] - p("hip_bone")[1])

    # cm conversions
    body_length_cm        = px_cm(body_length_px,       cmp)
    body_height_cm        = px_cm(body_height_px,        cmp)
    bbox_height_cm       = px_cm(bbox_height_px,        cmp)   # bbox height  (card + logs)
    hip_length_cm         = px_cm(hip_length_px,         cmp)
    chest_height_cm       = px_cm(chest_height_px,       cmp)
    linear_body_depth_cm  = px_cm(linear_body_depth_px,  cmp)
    rump_vertical_cm      = (rump_vertical_px * cmp) if (rump_vertical_px is not None and cmp) else None
    rump_vertical_cm      = rump_vertical_cm -5;
    # Derived
    heart_girth_cm  = (1.588 * chest_height_cm + 73.43) if chest_height_cm else None
    body_weight_kg  = ((heart_girth_cm**2 * body_length_cm) / 10840.0
                       if (heart_girth_cm and body_length_cm) else None)
    body_depth_cm   = (2.473 * linear_body_depth_cm + 53.015) if linear_body_depth_cm else None

    # Rear-leg set score (side view)
    rear_leg_set_score = None
    rear_leg_angle_deg = None
    if vis("hock") and vis("pastern"):
        dx = float(p("pastern")[0] - p("hock")[0])
        dy = float(p("pastern")[1] - p("hock")[1])
        angle_with_horiz = abs(math.degrees(math.atan2(abs(dy), abs(dx))))
        angle_with_horiz = min(angle_with_horiz, 180 - angle_with_horiz)
        rear_leg_angle_deg  = angle_with_horiz
        raw_score = 38.15 - 0.4384 * angle_with_horiz
        rear_leg_set_score = int(np.clip(round(raw_score), 1, 9))

    # Rump angle
    rump_angle_deg = None
    if vis("hip_bone") and vis("pin_bone") and hip_length_px:
        rump_angle_deg = abs(math.degrees(
            math.atan2(
                abs(float(p("pin_bone")[1] - p("hip_bone")[1])),
                abs(float(p("pin_bone")[0] - p("hip_bone")[0]))
            )
        ))

    # ── NEW: Foot angle ──────────────────────────────────────────────────────
    # Acute angle at vertex heel_bulb:
    #   arms → pastern_hoof_junction and hoof_tip
    foot_angle_deg   = None
    foot_angle_score = None
    if vis("pastern_hoof_junction") and vis("heel_bulb") and vis("hoof_tip"):
        a = p("pastern_hoof_junction")
        b = p("heel_bulb")          # vertex
        c = p("hoof_tip")
        angle = angle_at_vertex(a, b, c)
        if angle is not None:
            # Ensure we always take the acute angle
            foot_angle_deg   = min(angle, 180.0 - angle)
            foot_angle_score = foot_angle_to_score(foot_angle_deg)

    return {
        "body_length_cm":       body_length_cm,
        "body_height_cm":       body_height_cm,
        "bbox_height_cm":      bbox_height_cm,
        "hip_length_cm":        hip_length_cm,
        "chest_height_cm":      chest_height_cm,
        "linear_body_depth_cm": linear_body_depth_cm,
        "body_depth_cm":        body_depth_cm,
        "rump_vertical_cm":     rump_vertical_cm,
        "rump_angle_deg":       rump_angle_deg,
        "heart_girth_cm":       heart_girth_cm,
        "body_weight_kg":       body_weight_kg,
        "rear_leg_angle_deg":   rear_leg_angle_deg,
        "rear_leg_set_score":   rear_leg_set_score,
        "foot_angle_deg":       foot_angle_deg,       # ← NEW
        "foot_angle_score":     foot_angle_score,     # ← NEW
    }

# ─────────────────────────────────────────────────────────────────────────────
# SIDE-VIEW ANNOTATION
# ─────────────────────────────────────────────────────────────────────────────
def draw_side_annotation(img, kps, m, cmp, tag, bbox=None):
    out  = img.copy()
    h, w = out.shape[:2]
    font = cv2.FONT_HERSHEY_SIMPLEX
    fs   = max(0.38, w/2200)
    tk   = max(1, w//700)

    def p(name):  return (int(kps[SKP[name]][0]), int(kps[SKP[name]][1]))
    def vis(name): return float(kps[SKP[name]][2]) > 0
    # ── Bounding box ─────────────────────────────────────────────────────────
    if bbox is not None:
        x1, y1, x2, y2 = int(bbox[0]), int(bbox[1]), int(bbox[2]), int(bbox[3])
        # Draw bbox — semi-transparent fill + solid border
        overlay = out.copy()
        cv2.rectangle(overlay, (x1, y1), (x2, y2), (0, 200, 100), -1)
        cv2.addWeighted(overlay, 0.07, out, 0.93, 0, out)
        cv2.rectangle(out, (x1, y1), (x2, y2), (0, 220, 110), max(2, tk))
        # Label with bbox height
        bh = m.get("bbox_height_cm")
        bbox_lbl = f"BBox H: {bh:.1f}cm" if bh else "BBox"
        cv2.putText(out, bbox_lbl, (x1 + 4, y1 - 6), font, fs * 0.9, (0, 0, 0),      tk + 1)
        cv2.putText(out, bbox_lbl, (x1 + 4, y1 - 6), font, fs * 0.9, (0, 220, 110),   tk)

    # Skeleton lines
    pairs = [
        ("shoulder_bone", "pin_bone",      (0,255,0),    "BL"),
        ("wither",        "foot",           (255,80,0),   "BH"),
        ("hip_bone",      "pin_bone",       (0,0,255),    "HL"),
        ("chest_top",     "chest_bottom",   (0,200,255),  "CH"),
        ("body_top",      "body_bottom",    (255,0,200),  "LBD"),
    ]
    for n1,n2,col,lbl in pairs:
        if vis(n1) and vis(n2):
            cv2.line(out, p(n1), p(n2), col, tk+1)
            mx,my = (p(n1)[0]+p(n2)[0])//2, (p(n1)[1]+p(n2)[1])//2
            val_key = {
                "BL":  "body_length_cm",
                "BH":  "body_height_cm",
                "HL":  "hip_length_cm",
                "CH":  "chest_height_cm",
                "LBD": "linear_body_depth_cm",
            }.get(lbl)
            val = m.get(val_key)
            txt = f"{lbl}:{val:.1f}cm" if val else lbl
            cv2.putText(out, txt, (mx+4, my-4), font, fs*0.85, (0,0,0),  tk+1)
            cv2.putText(out, txt, (mx+4, my-4), font, fs*0.85, col,       tk)

    # Rear-leg angle
    if vis("hock") and vis("pastern"):
        hk, ps = p("hock"), p("pastern")
        cv2.line(out, hk, ps, (255,255,0), tk+1)
        hl1 = (ps[0]-60, ps[1]); hl2 = (ps[0]+60, ps[1])
        cv2.line(out, hl1, hl2, (200,200,0), tk, cv2.LINE_AA)
        cv2.putText(out, f"{m.get('rear_leg_angle_deg',0):.1f}°",
                    (ps[0]+5, ps[1]-8), font, fs*0.9, (255,255,0), tk)

    # Rump vertical
    if vis("hip_bone") and vis("pin_bone"):
        hp, pb = p("hip_bone"), p("pin_bone")
        cv2.line(out, (hp[0], hp[1]), (hp[0], pb[1]), (200,200,255), tk, cv2.LINE_AA)
        rv = m.get("rump_vertical_cm")
        sign_str = "+" if rv is not None and rv >= 0 else ""
        rv_txt = f"RV:{sign_str}{rv:.1f}cm" if rv is not None else "RV:N/A"
        cv2.putText(out, rv_txt, (hp[0]+5, (hp[1]+pb[1])//2), font, fs*0.85, (0,0,0),      tk+1)
        cv2.putText(out, rv_txt, (hp[0]+5, (hp[1]+pb[1])//2), font, fs*0.85, (200,200,255), tk)

    # ── NEW: Foot angle triangle ─────────────────────────────────────────────
    if vis("pastern_hoof_junction") and vis("heel_bulb") and vis("hoof_tip"):
        phj = p("pastern_hoof_junction")
        hb  = p("heel_bulb")
        ht  = p("hoof_tip")
        # Draw the two arms of the angle
        cv2.line(out, hb, phj, (255,165,0), tk+1)   # heel_bulb → pastern_hoof_junction
        cv2.line(out, hb, ht,  (255,165,0), tk+1)   # heel_bulb → hoof_tip
        fa  = m.get("foot_angle_deg")
        fas = m.get("foot_angle_score")
        if fa is not None:
            fa_txt = f"FA:{fa:.1f}° S:{fas}"
            cv2.putText(out, fa_txt, (hb[0]+5, hb[1]-8), font, fs*0.9, (0,0,0),      tk+2)
            cv2.putText(out, fa_txt, (hb[0]+5, hb[1]-8), font, fs*0.9, (255,165,0),   tk)

    # Keypoints
    for i,(x,y,v) in enumerate(kps):
        if v>0:
            col = SIDE_KP_COLORS[i]
            cv2.circle(out,(int(x),int(y)),max(4,w//160),col,-1)
            cv2.circle(out,(int(x),int(y)),max(4,w//160)+1,(0,0,0),1)
            cv2.putText(out,SIDE_KP_NAMES[i],(int(x)+5,int(y)-5),font,fs*0.75,(255,255,255),tk)
            cv2.putText(out,SIDE_KP_NAMES[i],(int(x)+5,int(y)-5),font,fs*0.75,col,max(1,tk-1))

    # Info panel
    lh = int(fs*52+8)
    rv = m.get("rump_vertical_cm")
    sign_str = "+" if rv is not None and rv >= 0 else ""
    fa  = m.get("foot_angle_deg")
    fas = m.get("foot_angle_score")
    lines = [
        (f"Tag: {tag}",                                                (0,255,180)),
        (f"Scale: {cmp:.5f} cm/px",                                    (180,255,180)),
        ("",None),
        (f"BBox Height       : {m['bbox_height_cm']:.2f} cm"       if m.get('bbox_height_cm')        else "BBox Height       : N/A", (0,220,110)),
        (f"Body Length       : {m['body_length_cm']:.2f} cm"         if m['body_length_cm']        else "Body Length       : N/A", (0,255,0)),
        (f"Body Height       : {m['body_height_cm']:.2f} cm"         if m['body_height_cm']        else "Body Height       : N/A", (255,80,0)),
        (f"Linear Body Depth : {m['linear_body_depth_cm']:.2f} cm"   if m['linear_body_depth_cm']  else "Linear Body Depth : N/A", (255,0,200)),
        (f"Body Depth        : {m['body_depth_cm']:.2f} cm"          if m['body_depth_cm']         else "Body Depth        : N/A", (200,100,255)),
        (f"Hip Length        : {m['hip_length_cm']:.2f} cm"          if m['hip_length_cm']         else "Hip Length        : N/A", (0,0,255)),
        (f"Chest Height      : {m['chest_height_cm']:.2f} cm"        if m['chest_height_cm']       else "Chest Height      : N/A", (0,200,255)),
        (f"Heart Girth       : {m['heart_girth_cm']:.2f} cm"         if m['heart_girth_cm']        else "Heart Girth       : N/A", (255,128,0)),
        (f"Est. Weight       : {m['body_weight_kg']:.1f} kg"          if m['body_weight_kg']        else "Est. Weight       : N/A", (0,255,100)),
        (f"Rump Vertical     : {sign_str}{rv:.2f} cm"                 if rv is not None             else "Rump Vertical     : N/A", (200,200,255)),
        (f"Rear Leg Angle    : {m['rear_leg_angle_deg']:.1f}°"        if m['rear_leg_angle_deg']    else "Rear Leg Angle    : N/A", (255,255,0)),
        (f"Rear Leg Score    : {m['rear_leg_set_score']}"              if m['rear_leg_set_score'] is not None else "Rear Leg Score    : N/A", (200,255,100)),
        (f"Foot Angle        : {fa:.1f}°  Score: {fas}"               if fa is not None             else "Foot Angle        : N/A", (255,165,0)),
    ]
    y0 = 28
    for txt, col in lines:
        if txt=="": y0+=6; continue
        cv2.putText(out,txt,(10,y0),font,fs,(0,0,0),tk+1)
        cv2.putText(out,txt,(10,y0),font,fs,col or (200,200,200),tk)
        y0+=lh
    return out

# ─────────────────────────────────────────────────────────────────────────────
# REAR-VIEW INFERENCE
# ─────────────────────────────────────────────────────────────────────────────
def run_rear_inference(img_bgr, yolo_model, min_conf=0.1):
    results = yolo_model(img_bgr, verbose=False)
    best_kps = None; best_conf = -1.0
    for r in results:
        if r.keypoints is None or r.boxes is None: continue
        kps_data = r.keypoints.data; confs = r.boxes.conf
        for i in range(len(kps_data)):
            c = float(confs[i])
            if c > best_conf:
                best_conf = c
                best_kps  = kps_data[i].cpu().numpy()
    if best_kps is None: return None
    kps_dict = {}
    for i,name in enumerate(REAR_KP_NAMES):
        if i>=len(best_kps): break
        x,y,conf = best_kps[i]
        if conf > min_conf:
            kps_dict[name] = [float(x),float(y)]
    return kps_dict if kps_dict else None

def compute_rear_measurements(kps_dict, cmp):
    dists = {}
    for kp1,kp2,label in REAR_MEASUREMENTS:
        if kp1 in kps_dict and kp2 in kps_dict:
            px_d = dist2d(kps_dict[kp1], kps_dict[kp2])
            dists[label] = px_cm(px_d, cmp)
        else:
            dists[label] = None

    rear_udder_height_cm = dists.get("M09_uddertop-udderbottom")

    def hoof_ground_angle(thigh_k, hock_k, hoof_k):
        if all(k in kps_dict for k in (thigh_k, hock_k, hoof_k)):
            hock = kps_dict[hock_k]; hoof = kps_dict[hoof_k]
            dx = hoof[0]-hock[0]; dy = hoof[1]-hock[1]
            ang = abs(math.degrees(math.atan2(abs(dy), abs(dx))))
            return min(ang, 180-ang)
        return None

    left_hoof_angle  = hoof_ground_angle("left_thigh",  "left_hock",  "hoof_left")
    right_hoof_angle = hoof_ground_angle("right_thigh", "right_hock", "hoof_right")

    return dists, left_hoof_angle, right_hoof_angle, rear_udder_height_cm

def compute_rf_score(rf_model, hock_dist, left_angle, right_angle):
    if rf_model is None:
        return None
    if hock_dist is None or left_angle is None or right_angle is None:
        return None
    try:
        mean_a = (left_angle + right_angle) / 2.0
        diff_a = abs(left_angle - right_angle)
        X = np.array([[hock_dist, left_angle, right_angle, mean_a, diff_a]])
        raw = rf_model.predict(X)[0]
        return int(np.clip(round(float(raw)), 1, 7))
    except Exception:
        return None

# ─────────────────────────────────────────────────────────────────────────────
# REAR-VIEW ANNOTATION
# ─────────────────────────────────────────────────────────────────────────────
def draw_rear_annotation(img, kps_dict, dists, left_angle, right_angle,
                          rear_udder_height_cm, cmp, tag):
    out  = img.copy()
    h, w = out.shape[:2]
    font = cv2.FONT_HERSHEY_SIMPLEX
    fs   = max(0.35, w/2400)
    tk   = max(1, w//600)

    for n1,n2 in REAR_SKELETON:
        if n1 in kps_dict and n2 in kps_dict:
            p1=(int(kps_dict[n1][0]),int(kps_dict[n1][1]))
            p2=(int(kps_dict[n2][0]),int(kps_dict[n2][1]))
            cv2.line(out,p1,p2,(0,255,0),tk)

    if "uddertop" in kps_dict and "udderbottom" in kps_dict:
        ut=(int(kps_dict["uddertop"][0]),int(kps_dict["uddertop"][1]))
        ub=(int(kps_dict["udderbottom"][0]),int(kps_dict["udderbottom"][1]))
        cv2.line(out, ut, ub, (0,180,255), tk+2)
        mx,my = (ut[0]+ub[0])//2, (ut[1]+ub[1])//2
        txt_uh = f"UH:{rear_udder_height_cm:.1f}cm" if rear_udder_height_cm else "UH:N/A"
        cv2.putText(out, txt_uh, (mx+4, my-4), font, fs*0.9, (0,0,0),      tk+2)
        cv2.putText(out, txt_uh, (mx+4, my-4), font, fs*0.9, (0,180,255),   tk)

    for kp1,kp2,label in REAR_MEASUREMENTS:
        if kp1 in kps_dict and kp2 in kps_dict and dists.get(label):
            p1=kps_dict[kp1]; p2=kps_dict[kp2]
            mx,my=int((p1[0]+p2[0])/2),int((p1[1]+p2[1])/2)
            short=label.split("_")[0]
            txt=f'{short}:{dists[label]:.1f}cm'
            cv2.putText(out,txt,(mx+2,my-4),font,fs*0.78,(0,0,0),tk+1)
            cv2.putText(out,txt,(mx+2,my-4),font,fs*0.78,(255,255,0),tk)

    for hock_k, hoof_k, angle_val, acol in [
        ("left_hock",  "hoof_left",  left_angle,  (0,200,255)),
        ("right_hock", "hoof_right", right_angle, (255,128,0)),
    ]:
        if hock_k in kps_dict and hoof_k in kps_dict and angle_val is not None:
            hk=(int(kps_dict[hock_k][0]),int(kps_dict[hock_k][1]))
            hf=(int(kps_dict[hoof_k][0]),int(kps_dict[hoof_k][1]))
            cv2.line(out,hk,hf,acol,tk+1)
            hl1=(hf[0]-50,hf[1]); hl2=(hf[0]+50,hf[1])
            cv2.line(out,hl1,hl2,(150,150,150),tk,cv2.LINE_AA)
            cv2.putText(out,f"{angle_val:.1f}°",(hf[0]+5,hf[1]-8),font,fs*0.9,(0,0,0),tk+2)
            cv2.putText(out,f"{angle_val:.1f}°",(hf[0]+5,hf[1]-8),font,fs*0.9,acol,tk)

    for i,name in enumerate(REAR_KP_NAMES):
        if name not in kps_dict: continue
        x,y=int(kps_dict[name][0]),int(kps_dict[name][1])
        col=REAR_KP_COLORS[i]
        r=max(4,w//160)
        cv2.circle(out,(x,y),r,col,-1)
        cv2.circle(out,(x,y),r+1,(0,0,0),1)
        cv2.putText(out,name,(x+6,y-4),font,fs,(255,255,255),tk)
        cv2.putText(out,name,(x+6,y-4),font,fs,col,max(1,tk-1))

    lh = int(fs*52+8); y0=28
    n_valid=sum(1 for v in dists.values() if v is not None)
    header=f"Rear View | Tag:{tag} | {n_valid}/{len(dists)} measurements"
    cv2.putText(out,header,(10,y0),font,fs*1.0,(0,0,0),tk+2)
    cv2.putText(out,header,(10,y0),font,fs*1.0,(0,255,180),tk)
    y0+=lh+4
    for _,_,label in REAR_MEASUREMENTS:
        val=dists.get(label)
        txt=f"{label}: {val:.1f}cm" if val else f"{label}: N/A"
        cv2.putText(out,txt,(10,y0),font,fs*0.78,(0,0,0),tk+1)
        cv2.putText(out,txt,(10,y0),font,fs*0.78,(0,255,180),tk)
        y0+=lh
    txt_uh = f"Rear Udder Height: {rear_udder_height_cm:.2f}cm" if rear_udder_height_cm else "Rear Udder Height: N/A"
    cv2.putText(out,txt_uh,(10,y0),font,fs*0.9,(0,0,0),tk+2)
    cv2.putText(out,txt_uh,(10,y0),font,fs*0.9,(0,180,255),tk)
    y0+=lh
    for lbl,val,acol in [
        ("Left hoof angle" ,left_angle, (0,200,255)),
        ("Right hoof angle",right_angle,(255,128,0)),
    ]:
        txt=f"{lbl}: {val:.2f}°" if val else f"{lbl}: N/A"
        cv2.putText(out,txt,(10,y0),font,fs*0.9,(0,0,0),tk+2)
        cv2.putText(out,txt,(10,y0),font,fs*0.9,acol,tk)
        y0+=lh
    return out

# ─────────────────────────────────────────────────────────────────────────────
# LOG HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _to_python(obj):
    if isinstance(obj, dict):
        return {k: _to_python(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_to_python(v) for v in obj]
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj

def load_logs():
    if os.path.exists(LOG_FILE):
        try:
            with open(LOG_FILE,"r") as f:
                return json.load(f)
        except: pass
    return []

def append_log(entry):
    logs = load_logs()
    logs.append(_to_python(entry))
    with open(LOG_FILE,"w") as f:
        json.dump(logs, f, indent=2)

def save_excel_log(logs):
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    ctr      = Alignment(horizontal="center", vertical="center")
    thin     = Side(style="thin", color="B0B0B0")
    brd      = Border(left=thin,right=thin,top=thin,bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title="Cattle Measurements"

    headers = ["Tag","Date","Time",
               "Height_BBox(cm)",
               "Height(cm)","Body_Length(cm)",
               "Linear_Body_Depth(cm)",
               "Body_Depth(cm)",
               "Heart_Girth(cm)",
               "Hip_Length(cm)","Chest_Height(cm)","Body_Weight(kg)",
               "Rump_Vertical(cm)",
               "Rump_Angle(deg)",
               "Rear_Leg_Angle(deg)","Rear_Leg_Set_Score(1-9)",
               "Foot_Angle(deg)",          # ← NEW
               "Foot_Angle_Score(1-9)",    # ← NEW
               "Hock_Distance(cm)",
               "Left_Hoof_Angle(deg)","Right_Hoof_Angle(deg)",
               "Rear_Leg_Rear_Score(1-7)",
               "Rear_Udder_Height(cm)",
               "M01_cm","M02_cm","M03_cm","M04_cm","M05_cm","M06_cm","M07_cm",
               "M08_cm","M09_cm","M10_cm","M11_cm","M12_cm","M13_cm","M14_cm",
               "M15_cm","M16_cm","M17_cm","M18_cm","M19_cm","M20_cm",
               "Thigh_Width(cm)"]

    for ci,h in enumerate(headers,1):
        c=ws.cell(1,ci,h)
        c.fill=hdr_fill;c.font=hdr_font;c.alignment=ctr;c.border=brd
    ws.freeze_panes="A2"

    for ri,entry in enumerate(logs,2):
        s  = entry.get("side",{})
        r  = entry.get("rear",{})
        rd = r.get("dists",{})
        row_vals = [
            entry.get("tag",""),
            entry.get("date",""),
            entry.get("time",""),
            _fmt(s.get("bbox_height_cm")),
            _fmt(s.get("body_height_cm")),
            _fmt(s.get("body_length_cm")),
            _fmt(s.get("linear_body_depth_cm")),
            _fmt(s.get("body_depth_cm")),
            _fmt(s.get("heart_girth_cm")),
            _fmt(s.get("hip_length_cm")),
            _fmt(s.get("chest_height_cm")),
            _fmt(s.get("body_weight_kg")),
            _fmt(s.get("rump_vertical_cm")),
            _fmt(s.get("rump_angle_deg")),
            _fmt(s.get("rear_leg_angle_deg")),
            _fmt(s.get("rear_leg_set_score"), nd=0),
            _fmt(s.get("foot_angle_deg")),          # ← NEW
            _fmt(s.get("foot_angle_score"), nd=0),  # ← NEW
            _fmt(r.get("hock_dist_cm")),
            _fmt(r.get("left_hoof_angle")),
            _fmt(r.get("right_hoof_angle")),
            _fmt(r.get("rf_score"), nd=0),
            _fmt(r.get("rear_udder_height_cm")),
        ]
        m_labels=[lbl for _,_,lbl in REAR_MEASUREMENTS]
        for lbl in m_labels:
            row_vals.append(_fmt(rd.get(lbl)))

        for ci,val in enumerate(row_vals,1):
            c=ws.cell(ri,ci,val)
            c.alignment=ctr;c.border=brd
            if ri%2==0: c.fill=alt_fill

    for col_cells in ws.columns:
        cw=max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width=min(cw+4,30)
    wb.save(EXCEL_FILE)

def _fmt(v, nd=2):
    if v is None: return "N/A"
    try: return round(float(v), nd)
    except: return v

# ─────────────────────────────────────────────────────────────────────────────
# IMAGE DISPLAY HELPER
# ─────────────────────────────────────────────────────────────────────────────
def bgr_to_pil(img_bgr):
    return Image.fromarray(cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB))

def resize_for_display(img_bgr, max_w=620, max_h=460):
    h, w = img_bgr.shape[:2]
    scale = min(max_w / w, max_h / h, 1.0)
    if scale < 1.0:
        new_w, new_h = int(w * scale), int(h * scale)
        img_bgr = cv2.resize(img_bgr, (new_w, new_h), interpolation=cv2.INTER_AREA)
    return img_bgr

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR — Navigation with auto-close on mobile
# ─────────────────────────────────────────────────────────────────────────────
NAV_OPTIONS = ["Measure", "Output", "Logs", "Settings"]

def _on_nav_change():
    st.session_state.nav = st.session_state._nav_radio

with st.sidebar:
    st.markdown("## 🐄 CattleScan")
    st.markdown("---")

    st.radio(
        "Navigation",
        NAV_OPTIONS,
        index=NAV_OPTIONS.index(st.session_state.nav),
        key="_nav_radio",
        on_change=_on_nav_change,
    )
    st.markdown("---")

    st.markdown("### 📁 Model Files")
    side_ckpt  = st.text_input("Side-view checkpoint (.pth)", value="Resnet101_Rcnn_Model.pth")
    rear_model = st.text_input("Rear-view YOLO model (.pt)",  value="rear_best_v2.pt")
    sam2_path  = st.text_input("SAM2 weights (.pt)",          value="sam2_b.pt")
    rf_pkl     = st.text_input("RF score model (.pkl)",       value="rf_score_model.pkl")
    st.markdown("---")
    st.markdown("### ⚙️ Settings")
    sticker_cm = st.number_input("Sticker real size (cm)", value=15.0, step=0.5)
    STICKER_REAL_CM = sticker_cm
    min_conf   = st.slider("YOLO min keypoint conf", 0.05, 0.5, 0.1, 0.01)
    st.markdown("---")
    st.caption("v1.2  •  Cattle Biometrics Lab")

# ── Auto-close sidebar on mobile after nav click ─────────────────────────────
st.markdown("""
<script>
(function() {
    function tryCloseSidebar() {
        if (window.innerWidth > 768) return;
        var btns = document.querySelectorAll(
            '[data-testid="stSidebarCollapseButton"], [aria-label="Close sidebar"], [aria-label="collapse sidebar"]'
        );
        btns.forEach(function(btn) { btn.click(); });
    }
    // Listen for radio changes inside sidebar
    document.addEventListener('change', function(e) {
        if (e.target && e.target.type === 'radio' &&
            e.target.closest('[data-testid="stSidebar"]')) {
            setTimeout(tryCloseSidebar, 300);
        }
    });
    document.addEventListener('click', function(e) {
        var label = e.target.closest('[data-testid="stSidebar"] label');
        if (label) { setTimeout(tryCloseSidebar, 300); }
    });
})();
</script>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div class="header-icon">🐄</div>
  <div>
    <h1>CattleScan</h1>
  </div>
</div>
""", unsafe_allow_html=True)

# helper: render one trait card
def _tc(label, value, unit, fmt=".1f"):
    if value is not None:
        try:    vstr = format(float(value), fmt)
        except: vstr = str(value)
    else:
        vstr = "—"
    return f"""<div class="trait-card">
  <div class="tc-label">{label}</div>
  <div class="tc-value">{vstr}</div>
  <div class="tc-unit">{unit}</div>
</div>"""

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: MEASURE
# ═════════════════════════════════════════════════════════════════════════════
if st.session_state.nav == "Measure":

    col_tag, col_btn = st.columns([3,1])
    with col_tag:
        tag_input = st.text_input("🏷️  Cattle Tag / ID",
                                   value=st.session_state.tag,
                                   placeholder="e.g.  COW_001")
        st.session_state.tag = tag_input.strip()
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄  New Cattle", use_container_width=True):
            keys_to_clear = [k for k in st.session_state if k not in ("nav", "_nav_radio")]
            for k in keys_to_clear:
                del st.session_state[k]
            init_state()
            st.rerun()

    if not st.session_state.tag:
        st.markdown('<div class="warn-box">⚠️  Please enter a cattle tag before uploading images.</div>',
                    unsafe_allow_html=True)
        st.stop()

    st.markdown("---")
    _needs_rerun = False

    col_side, col_rear = st.columns(2, gap="large")

    # ══════════════════════════════════════════
    # SIDE VIEW
    # ══════════════════════════════════════════
    with col_side:
        st.markdown('<div class="section-head">📷  Side View</div>', unsafe_allow_html=True)

        side_src = st.radio("Image source", ["Upload","Camera"], key="side_src", horizontal=True)
        if side_src == "Upload":
            side_file = st.file_uploader("Upload side-view image",
                                          type=["jpg","jpeg","png","bmp"],
                                          key="side_uploader")
            if side_file is not None:
                file_id = (side_file.name, side_file.size)
                if st.session_state.get("side_file_id") != file_id:
                    arr = np.frombuffer(side_file.read(), np.uint8)
                    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                    st.session_state.side_img_raw              = img
                    st.session_state.side_sticker_done         = False
                    st.session_state.side_results              = None
                    st.session_state.side_last_processed_click = None
                    st.session_state.side_file_id              = file_id
        else:
            cam_img = st.camera_input("Capture side-view image", key="side_cam")
            if cam_img is not None:
                cam_id = cam_img.file_id if hasattr(cam_img, "file_id") else id(cam_img)
                if st.session_state.get("side_file_id") != cam_id:
                    arr = np.frombuffer(cam_img.read(), np.uint8)
                    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                    st.session_state.side_img_raw              = img
                    st.session_state.side_sticker_done         = False
                    st.session_state.side_results              = None
                    st.session_state.side_last_processed_click = None
                    st.session_state.side_file_id              = cam_id

        if st.session_state.side_img_raw is not None:
            img_raw  = st.session_state.side_img_raw
            disp     = resize_for_display(img_raw)
            pil_disp = bgr_to_pil(disp)

            if not st.session_state.side_sticker_done:
                st.markdown('<div class="sticker-box">👆  Click on the <b>square QR sticker</b> in the image below to calibrate scale</div>',
                            unsafe_allow_html=True)
                if COORDS_OK:
                    coords = streamlit_image_coordinates(pil_disp, key="side_click")
                    if coords:
                        sx = img_raw.shape[1] / disp.shape[1]
                        sy = img_raw.shape[0] / disp.shape[0]
                        orig_xy = (int(coords["x"] * sx), int(coords["y"] * sy))
                        if st.session_state.side_last_processed_click != orig_xy:
                            st.session_state.side_last_processed_click = orig_xy
                            with st.spinner("Segmenting sticker with SAM2…"):
                                sam2 = load_sam2(sam2_path) if model_file_available(sam2_path) else None
                                if sam2:
                                    cmp, _ = segment_sticker_sam2(img_raw, orig_xy, sam2)
                                else:
                                    cmp, _ = fallback_cm_per_px(img_raw, orig_xy)
                            if cmp:
                                st.session_state.side_cm_per_px    = cmp
                                st.session_state.side_sticker_done = True
                                _needs_rerun = True
                            else:
                                st.error("Sticker detection failed — try clicking the centre of the sticker.")
                else:
                    # Fallback: show image and manual scale entry
                    # FIX: Use st.image with explicit format to ensure visibility in all themes
                    img_bytes = io.BytesIO()
                    pil_disp.save(img_bytes, format="PNG")
                    st.image(img_bytes.getvalue(), use_container_width=True)
                    st.markdown('<div class="warn-box">Install <code>streamlit-image-coordinates</code> for click detection.</div>',
                                unsafe_allow_html=True)
                    manual_cmp = st.number_input("Enter scale manually (cm/px)", value=0.05,
                                                  format="%.5f", key="side_manual_cmp")
                    if st.button("✅ Use this scale (Side)", key="side_manual_btn"):
                        st.session_state.side_cm_per_px    = manual_cmp
                        st.session_state.side_sticker_done = True
                        _needs_rerun = True
            else:
                cmp_val = st.session_state.side_cm_per_px
                # FIX: Convert to PNG bytes for reliable display in all themes/modes
                img_bytes = io.BytesIO()
                pil_disp.save(img_bytes, format="PNG")
                st.image(img_bytes.getvalue(),
                         caption=f"✅ Scale: {cmp_val:.5f} cm/px",
                         use_container_width=True)
                if st.session_state.side_results:
                    st.success("✅ Side-view measurements complete")
                else:
                    st.info("Scale set. Click 'Estimate Trait Measurements' below.")

    # ══════════════════════════════════════════
    # REAR VIEW
    # ══════════════════════════════════════════
    with col_rear:
        st.markdown('<div class="section-head">📷  Rear View</div>', unsafe_allow_html=True)

        rear_src = st.radio("Image source", ["Upload","Camera"], key="rear_src", horizontal=True)
        if rear_src == "Upload":
            rear_file = st.file_uploader("Upload rear-view image",
                                          type=["jpg","jpeg","png","bmp"],
                                          key="rear_uploader")
            if rear_file is not None:
                file_id = (rear_file.name, rear_file.size)
                if st.session_state.get("rear_file_id") != file_id:
                    arr = np.frombuffer(rear_file.read(), np.uint8)
                    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                    st.session_state.rear_img_raw              = img
                    st.session_state.rear_sticker_done         = False
                    st.session_state.rear_results              = None
                    st.session_state.rear_last_processed_click = None
                    st.session_state.rear_file_id              = file_id
        else:
            cam_img = st.camera_input("Capture rear-view image", key="rear_cam")
            if cam_img is not None:
                cam_id = cam_img.file_id if hasattr(cam_img, "file_id") else id(cam_img)
                if st.session_state.get("rear_file_id") != cam_id:
                    arr = np.frombuffer(cam_img.read(), np.uint8)
                    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
                    st.session_state.rear_img_raw              = img
                    st.session_state.rear_sticker_done         = False
                    st.session_state.rear_results              = None
                    st.session_state.rear_last_processed_click = None
                    st.session_state.rear_file_id              = cam_id

        if st.session_state.rear_img_raw is not None:
            img_raw  = st.session_state.rear_img_raw
            disp     = resize_for_display(img_raw)
            pil_disp = bgr_to_pil(disp)

            if not st.session_state.rear_sticker_done:
                st.markdown('<div class="sticker-box">👆  Click on the <b>square QR sticker</b> in the image below to calibrate scale</div>',
                            unsafe_allow_html=True)
                if COORDS_OK:
                    coords = streamlit_image_coordinates(pil_disp, key="rear_click")
                    if coords:
                        sx = img_raw.shape[1] / disp.shape[1]
                        sy = img_raw.shape[0] / disp.shape[0]
                        orig_xy = (int(coords["x"] * sx), int(coords["y"] * sy))
                        if st.session_state.rear_last_processed_click != orig_xy:
                            st.session_state.rear_last_processed_click = orig_xy
                            with st.spinner("Segmenting sticker with SAM2…"):
                                sam2 = load_sam2(sam2_path) if model_file_available(sam2_path) else None
                                if sam2:
                                    cmp, _ = segment_sticker_sam2(img_raw, orig_xy, sam2)
                                else:
                                    cmp, _ = fallback_cm_per_px(img_raw, orig_xy)
                            if cmp:
                                st.session_state.rear_cm_per_px    = cmp
                                st.session_state.rear_sticker_done = True
                                _needs_rerun = True
                            else:
                                st.error("Sticker detection failed — try clicking the centre of the sticker.")
                else:
                    img_bytes = io.BytesIO()
                    pil_disp.save(img_bytes, format="PNG")
                    st.image(img_bytes.getvalue(), use_container_width=True)
                    manual_cmp = st.number_input("Enter scale manually (cm/px)", value=0.05,
                                                  format="%.5f", key="rear_manual_cmp")
                    if st.button("✅ Use this scale (Rear)", key="rear_manual_btn"):
                        st.session_state.rear_cm_per_px    = manual_cmp
                        st.session_state.rear_sticker_done = True
                        _needs_rerun = True
            else:
                cmp_val = st.session_state.rear_cm_per_px
                img_bytes = io.BytesIO()
                pil_disp.save(img_bytes, format="PNG")
                st.image(img_bytes.getvalue(),
                         caption=f"✅ Scale: {cmp_val:.5f} cm/px",
                         use_container_width=True)
                if st.session_state.rear_results:
                    st.success("✅ Rear-view measurements complete")
                else:
                    st.info("Scale set. Click 'Estimate Trait Measurements' below.")

    if _needs_rerun:
        st.rerun()

    # ── ESTIMATE button ───────────────────────────────────────────────────────
    st.markdown("---")
    side_ready = (st.session_state.side_img_raw is not None and
                  st.session_state.side_sticker_done)
    rear_ready = (st.session_state.rear_img_raw is not None and
                  st.session_state.rear_sticker_done)

    btn_col = st.columns([1, 2, 1])[1]
    with btn_col:
        estimate_btn = st.button("🔬  Estimate Trait Measurements",
                                  disabled=not (side_ready or rear_ready),
                                  use_container_width=True,
                                  type="primary")

    if estimate_btn:
        side_m = {}
        rear_m = {}

        if side_ready:
            with st.spinner("Running side-view keypoint detection…"):
                try:
                    if TORCH_OK and model_file_available(side_ckpt):
                        result = load_side_model(side_ckpt)
                        if result is None:
                            st.warning("Side-view model failed to load.")
                        else:
                            model, device = result
                            kps, det_score, bbox = run_side_inference(
                                st.session_state.side_img_raw, model, device)
                            if kps is not None:
                                cmp    = st.session_state.side_cm_per_px
                                side_m = compute_side_measurements(kps, cmp, bbox=bbox)
                                ann    = draw_side_annotation(
                                    st.session_state.side_img_raw, kps, side_m,
                                    cmp, st.session_state.tag, bbox=bbox)
                                st.session_state.side_annotated = ann
                                st.session_state.side_results   = side_m
                                cv2.imwrite(str(OUTPUT_DIR / "side_annotated.jpg"), ann,
                                            [cv2.IMWRITE_JPEG_QUALITY, 95])
                            else:
                                st.warning("No cattle detected in side-view image.")
                    elif not TORCH_OK:
                        st.error(
                            "Side-view inference is unavailable because PyTorch "
                            f"or torchvision failed to import: {TORCH_IMPORT_ERROR}"
                        )
                    else:
                        st.warning(
                            "Side-view model is missing and no Google Drive URL "
                            "is configured — skipping side-view."
                        )
                except Exception as e:
                    st.error(f"Side-view error: {e}")
                    import traceback; st.code(traceback.format_exc())

        if rear_ready:
            with st.spinner("Running rear-view keypoint detection…"):
                try:
                    if YOLO_OK and model_file_available(rear_model):
                        yolo     = load_rear_model(rear_model)
                        if yolo is None:
                            st.warning("Rear-view model failed to load.")
                        else:
                            kps_dict = run_rear_inference(
                                st.session_state.rear_img_raw, yolo, min_conf)
                            if kps_dict is not None:
                                cmp       = st.session_state.rear_cm_per_px
                                dists, la, ra, udder_ht = compute_rear_measurements(kps_dict, cmp)
                                rf_score  = None
                                hock_dist = dists.get("M20_left_hock-right_hock")
                                if model_file_available(rf_pkl):
                                    try:
                                        rf = load_rf_model(rf_pkl)
                                    except Exception:
                                        rf = None
                                    rf_score = compute_rf_score(rf, hock_dist, la, ra)
                                rear_m = {
                                    "dists":                dists,
                                    "left_hoof_angle":      la,
                                    "right_hoof_angle":     ra,
                                    "hock_dist_cm":         hock_dist,
                                    "rf_score":             rf_score,
                                    "rear_udder_height_cm": udder_ht,
                                }
                                ann = draw_rear_annotation(
                                    st.session_state.rear_img_raw, kps_dict,
                                    dists, la, ra, udder_ht,
                                    cmp, st.session_state.tag)
                                st.session_state.rear_annotated = ann
                                st.session_state.rear_results   = rear_m
                                cv2.imwrite(str(OUTPUT_DIR / "rear_annotated.jpg"), ann,
                                            [cv2.IMWRITE_JPEG_QUALITY, 95])
                            else:
                                st.warning("No cattle detected in rear-view image.")
                    else:
                        st.warning("Rear-view model not found — skipping rear-view.")
                except Exception as e:
                    st.error(f"Rear-view error: {e}")
                    import traceback; st.code(traceback.format_exc())

        if side_m or rear_m:
            now = datetime.now()
            log_entry = {
                "tag":  st.session_state.tag,
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%H:%M:%S"),
                "side": side_m,
                "rear": rear_m,
            }
            st.session_state.traits = log_entry
            append_log(log_entry)
            save_excel_log(load_logs())

        st.rerun()

    # ── DISPLAY TRAITS ────────────────────────────────────────────────────────
    if st.session_state.traits:
        t  = st.session_state.traits
        sm = t.get("side", {})
        rm = t.get("rear",  {})
        rd = rm.get("dists", {})

        st.markdown("---")
        st.markdown(f'<div class="section-head">📊  Estimated Traits — <span style="color:#40916c">{t["tag"]}</span></div>',
                    unsafe_allow_html=True)

        # ── Build 10-trait responsive grid via HTML ───────────────────────────
        # Rump vertical (signed display)
        rv = sm.get("rump_vertical_cm")
        rv_str = (f"{'+' if rv >= 0 else ''}{rv:.1f}") if rv is not None else "—"

        # Rear leg set score
        rls = sm.get("rear_leg_set_score")
        rls_str = str(rls) if rls is not None else "—"

        # Rear leg rear view (RF score)
        rf_s = rm.get("rf_score")
        rf_str = str(rf_s) if rf_s is not None else "—"

        # Rump width (M03)
        rump_w = rd.get("M03_rump_left-rump_right")

        # Foot angle & score
        fa  = sm.get("foot_angle_deg")
        fas = sm.get("foot_angle_score")
        fa_str  = f"{fa:.1f}" if fa is not None else "—"
        fas_str = str(fas) if fas is not None else "—"

        # 10 cards — fully self-contained HTML+CSS so Streamlit can't strip styles
        def make_card(label, value_html, unit_html):
            return f"""
            <div style="
                background: linear-gradient(135deg,#1a3a5c,#1e4d6b);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 14px;
                padding: 16px 12px;
                text-align: center;
                box-shadow: 0 2px 12px rgba(0,0,0,0.15);
                min-height: 90px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
            ">
              <div style="color:#8ecae6;font-size:0.72rem;font-weight:600;
                          text-transform:uppercase;letter-spacing:0.06em;
                          margin-bottom:6px;line-height:1.3;">{label}</div>
              <div style="color:#ffffff;font-size:1.45rem;font-weight:700;
                          line-height:1.2;">{value_html}</div>
              <div style="color:#90caf9;font-size:0.76rem;margin-top:4px;">{unit_html}</div>
            </div>"""

        def simple_card(label, value, unit, fmt=".1f"):
            if value is not None:
                try:    vstr = format(float(value), fmt)
                except: vstr = str(value)
            else:
                vstr = "—"
            return make_card(label, vstr, unit)

        # 10 traits in a responsive CSS grid
        cards_html = (
            _tc("Height",             sm.get("bbox_height_cm"),           "cm") +
            _tc("Body Length",        sm.get("body_length_cm"),           "cm") +
            _tc("Body Depth",         sm.get("body_depth_cm"),            "cm") +
            _tc("Heart Girth",        sm.get("heart_girth_cm"),           "cm") +
            f"""<div class="trait-card">
              <div class="tc-label">Rear Leg Set</div>
              <div class="tc-value">{rls_str}</div>
              <div class="tc-unit">1–9 scale</div>
            </div>""" +
            f"""<div class="trait-card">
              <div class="tc-label">Rear Leg Rear View</div>
              <div class="tc-value">{rf_str}</div>
              <div class="tc-unit">1–9 scale</div>
            </div>""" +
            f"""<div class="trait-card">
              <div class="tc-label">Rump Angle</div>
              <div class="tc-value">{rv_str}</div>
              <div class="tc-unit">cm (signed)</div>
            </div>""" +
            _tc("Rump Width",         rd.get("M03_rump_left-rump_right"), "cm") +
            _tc("Rear Udder Height",  rm.get("rear_udder_height_cm"),     "cm") +
            f"""<div class="trait-card">
              <div class="tc-label">Foot Angle Score</div>
              <div class="tc-value">{fas_str}</div>
              <div class="tc-unit">1–9 scale</div>
            </div>"""
        )
        st.markdown(f'<div class="trait-grid">{cards_html}</div>', unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: OUTPUT
# ═════════════════════════════════════════════════════════════════════════════
elif st.session_state.nav == "Output":
    st.markdown('<div class="section-head">🖼️  Annotated Output — Current Cattle</div>',
                unsafe_allow_html=True)

    tag = st.session_state.get("tag","—")
    st.markdown(f"**Cattle:** `{tag}`")

    c1, c2 = st.columns(2)
    side_ann = st.session_state.get("side_annotated")
    rear_ann = st.session_state.get("rear_annotated")

    with c1:
        st.markdown("**Side View**")
        if side_ann is not None:
            # FIX: Use PNG bytes for reliable cross-theme display
            pil_side = bgr_to_pil(side_ann)
            buf = io.BytesIO()
            pil_side.save(buf, format="PNG")
            st.image(buf.getvalue(), use_container_width=True)
            buf_dl = io.BytesIO()
            pil_side.save(buf_dl, format="JPEG", quality=95)
            st.download_button("⬇️  Download Side Annotated",
                               buf_dl.getvalue(), "side_annotated.jpg",
                               mime="image/jpeg")
        else:
            st.info("No side-view output yet. Run measurements first.")

    with c2:
        st.markdown("**Rear View**")
        if rear_ann is not None:
            pil_rear = bgr_to_pil(rear_ann)
            buf = io.BytesIO()
            pil_rear.save(buf, format="PNG")
            st.image(buf.getvalue(), use_container_width=True)
            buf_dl = io.BytesIO()
            pil_rear.save(buf_dl, format="JPEG", quality=95)
            st.download_button("⬇️  Download Rear Annotated",
                               buf_dl.getvalue(), "rear_annotated.jpg",
                               mime="image/jpeg")
        else:
            st.info("No rear-view output yet. Run measurements first.")

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: LOGS
# ═════════════════════════════════════════════════════════════════════════════
elif st.session_state.nav == "Logs":
    st.markdown('<div class="section-head">📋  Measurement Logs — All Cattle</div>',
                unsafe_allow_html=True)

    logs = load_logs()
    if not logs:
        st.info("No measurements logged yet.")
    else:
        col1, col2 = st.columns([3,1])
        with col2:
            if os.path.exists(EXCEL_FILE):
                with open(EXCEL_FILE,"rb") as f:
                    st.download_button("⬇️  Download Excel", f.read(),
                                       "cattle_logs.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with col1:
            st.markdown(f"**{len(logs)} record(s) in log**")

        for entry in reversed(logs):
            sm = entry.get("side",{})
            rm = entry.get("rear",{})
            rd = rm.get("dists",{})

            with st.expander(f"🐄  {entry['tag']}  |  {entry['date']}  {entry['time']}", expanded=False):
                sc1, sc2 = st.columns(2)
                with sc1:
                    st.markdown("**Side-view measurements**")
                    rv = sm.get("rump_vertical_cm")
                    rv_disp = (f"{'+' if rv >= 0 else ''}{rv:.2f}" if rv is not None else "N/A")
                    fa  = sm.get("foot_angle_deg")
                    fas = sm.get("foot_angle_score")
                    side_rows = [
                        ("Height (BBox)",        sm.get("bbox_height_cm"),        "cm"),
                        ("Body Height (W→F)",    sm.get("body_height_cm"),        "cm"),
                        ("Body Length",          sm.get("body_length_cm"),         "cm"),
                        ("Linear Body Depth",    sm.get("linear_body_depth_cm"),   "cm"),
                        ("Body Depth",           sm.get("body_depth_cm"),          "cm"),
                        ("Hip Length",           sm.get("hip_length_cm"),          "cm"),
                        ("Chest Height",         sm.get("chest_height_cm"),        "cm"),
                        ("Heart Girth",          sm.get("heart_girth_cm"),         "cm"),
                        ("Est. Body Weight",     sm.get("body_weight_kg"),         "kg"),
                        ("Rump Angle",           None,                             "cm"),   # handled below
                        ("Rear Leg Angle",       sm.get("rear_leg_angle_deg"),     "°"),
                        ("Rear Leg Set Score",   sm.get("rear_leg_set_score"),     "/ 9"),
                        ("Foot Angle",           fa,                               "°"),    # ← NEW
                        ("Foot Angle Score",     fas,                              "/ 9"),  # ← NEW
                    ]
                    rows_html = ""
                    for lbl, v, u in side_rows:
                        if lbl == "Rump Angle":
                            rows_html += f"<tr><td>{lbl}</td><td>{rv_disp} {u}</td></tr>"
                        elif lbl == "Rear Leg Set Score":
                            val_str = str(int(v)) if v is not None else "N/A"
                            rows_html += f"<tr><td>{lbl}</td><td>{val_str} {u}</td></tr>"
                        elif lbl == "Foot Angle Score":
                            val_str = str(int(v)) if v is not None else "N/A"
                            rows_html += f"<tr><td>{lbl}</td><td>{val_str} {u}</td></tr>"
                        else:
                            rows_html += (
                                f"<tr><td>{lbl}</td>"
                                f"<td>{'%.2f' % v if v is not None else 'N/A'} {u}</td></tr>"
                            )
                    st.markdown(
                        f'<table class="log-table"><thead><tr><th>Measurement</th>'
                        f'<th>Value</th></tr></thead><tbody>{rows_html}</tbody></table>',
                        unsafe_allow_html=True)

                with sc2:
                    st.markdown("**Rear-view measurements**")
                    rear_rows = [(lbl, rd.get(lbl), "cm") for _,_,lbl in REAR_MEASUREMENTS]
                    rear_rows += [
                        ("Rear Udder Height", rm.get("rear_udder_height_cm"), "cm"),
                        ("Left Hoof Angle",   rm.get("left_hoof_angle"),      "°"),
                        ("Right Hoof Angle",  rm.get("right_hoof_angle"),     "°"),
                        ("Hock Distance",     rm.get("hock_dist_cm"),         "cm"),
                        ("RF Rear Score",     rm.get("rf_score"),             "/ 7"),
                    ]
                    rows_html = "".join(
                        f"<tr><td>{lbl}</td><td>"
                        f"{'%.2f' % v if isinstance(v,(int,float)) and v is not None else ('N/A' if v is None else v)}"
                        f" {u}</td></tr>"
                        for lbl,v,u in rear_rows
                    )
                    st.markdown(
                        f'<table class="log-table"><thead><tr><th>Measurement</th>'
                        f'<th>Value</th></tr></thead><tbody>{rows_html}</tbody></table>',
                        unsafe_allow_html=True)

# ═════════════════════════════════════════════════════════════════════════════
# PAGE: SETTINGS
# ═════════════════════════════════════════════════════════════════════════════
elif st.session_state.nav == "Settings":
    st.markdown('<div class="section-head">⚙️  Settings & Help</div>',
                unsafe_allow_html=True)

    st.markdown("""
    ### Required files
    Place the following files in the **same directory** as `app.py`:

    | File | Description |
    |------|-------------|
    | `Resnet101_Rcnn_Model.pth` | Side-view KeypointRCNN checkpoint |
    | `rear_best_v2.pt`           | Rear-view YOLOv8-pose weights |
    | `sam2_b.pt`                 | SAM2 segmentation weights |
    | `rf_score_model.pkl`        | Random Forest rear-leg-rear score model |

    ### Installation
    ```bash
    pip install streamlit streamlit-image-coordinates opencv-python-headless \\
                torch torchvision ultralytics openpyxl scikit-learn pillow
    ```

    ### Run
    ```bash
    streamlit run app.py
    ```

    ### Measurement pipeline
    1. Enter cattle tag.
    2. Upload or capture **side-view** image → click on the QR sticker.
    3. Upload or capture **rear-view** image → click on the QR sticker.
    4. Click **Estimate Trait Measurements**.
    5. View key traits on the **Measure** page.
    6. View annotated images on the **Output** page.
    7. All records (appended, never overwritten) in **Logs**.

    ### Field definitions & equations

    | Field | Description / Formula |
    |-------|-----------------------|
    | Linear Body Depth | Direct pixel distance between `body_top` and `body_bottom` keypoints (cm) |
    | **Body Depth** | Derived score: `2.473 × linear_body_depth + 53.015` |
    | Heart Girth | `1.588 × chest_height + 73.43` |
    | Body Weight | `(heart_girth² × body_length) / 10840` |
    | Rear Leg Set Score | `round(38.15 − 0.4384 × hock_pastern_angle)` clipped to 1–9 (integer) |
    | **Foot Angle** | Acute angle at `heel_bulb` between `pastern_hoof_junction` and `hoof_tip` (degrees) |
    | **Foot Angle Score** | 70–76°=1, 63–69°=2, 56–62°=3, 49–55°=4, 42–48°=5, 35–41°=6, 28–34°=7, 21–27°=8, 14–20°=9 |
    | Rear Udder Height | Distance between `uddertop` and `udderbottom` keypoints (cm) |
    | Rump Vertical | Signed vertical distance (cm): positive = hip_bone above pin_bone (normal slope) |
    | Rear Leg Rear Score | Random Forest model (1–7) |
    """)

    st.markdown("---")
    if st.button("🗑️  Clear ALL logs (irreversible)", type="secondary"):
        if os.path.exists(LOG_FILE):   os.remove(LOG_FILE)
        if os.path.exists(EXCEL_FILE): os.remove(EXCEL_FILE)
        st.success("Logs cleared.")
